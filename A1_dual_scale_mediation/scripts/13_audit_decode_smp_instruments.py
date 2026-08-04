#!/usr/bin/env python3
"""Audit deCODE clumped instruments for PAV/epitope and cross-assay reuse."""

from __future__ import annotations

import gzip
import hashlib
import os
import subprocess
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = ROOT.parent
WORKBOOK = ROOT / "data_raw" / "decode_public" / "41586_2023_6563_MOESM3_ESM.xlsx"
INSTRUMENTS = ROOT / "data_processed" / "decode_smp_clumped_instruments.tsv"
RESOURCE_ROOT = Path(os.environ.get("A1_RESOURCE_ROOT", PROJECT_ROOT / "data" / "external"))
PLINK = Path(os.environ.get("PLINK_BIN", "plink"))
LD_REFERENCE = Path(os.environ.get("A1_LD_PREFIX", str(RESOURCE_ROOT / "EUR" / "EUR")))
WORK_DIR = ROOT / "data_processed" / "decode_smp_shared_instrument_ld_audit"

PAV_OUTPUT = ROOT / "tables" / "decode_smp_PAV_epitope_instrument_audit.tsv"
SHARED_OUTPUT = ROOT / "tables" / "decode_smp_shared_instrument_audit.tsv"
ASSAY_OUTPUT = ROOT / "tables" / "decode_smp_assay_annotation_audit.tsv"
LOG_OUTPUT = ROOT / "logs" / "decode_smp_instrument_audit.log"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def clean_header(value: object) -> str:
    return str(value).replace("\n", "_").replace(" ", "_").strip("_") if value is not None else ""


def truthy(value: object) -> bool | None:
    if value is None or str(value).upper() in {"NA", "NONE", "NAN", ""}:
        return None
    return str(value).strip().upper() in {"Y", "YES", "TRUE", "1"}


def rs_aliases(*values: object) -> set[str]:
    aliases: set[str] = set()
    for value in values:
        if value is None:
            continue
        for item in str(value).split(","):
            item = item.strip()
            if item.startswith("rs"):
                aliases.add(item)
    return aliases


def read_target_supplement(seqids: set[str]) -> tuple[pd.DataFrame, list[dict]]:
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)

    assay_sheet = workbook["ST29_protein_classification"]
    assay_headers = [clean_header(value) for value in next(assay_sheet.iter_rows(min_row=3, max_row=3, values_only=True))]
    assay_rows = []
    for values in assay_sheet.iter_rows(min_row=4, values_only=True):
        row = dict(zip(assay_headers, values))
        if str(row.get("SeqId")) in seqids:
            assay_rows.append(row)

    locus_sheet = workbook["ST19_somascan_pQTLs_smpnorm"]
    locus_headers = [clean_header(value) for value in next(locus_sheet.iter_rows(min_row=4, max_row=4, values_only=True))]
    locus_rows = []
    for values in locus_sheet.iter_rows(min_row=5, values_only=True):
        row = dict(zip(locus_headers, values))
        if str(row.get("SeqId")) in seqids:
            row["rs_aliases"] = rs_aliases(row.get("variant"), row.get("LD_class"), row.get("Reported_rsid"))
            locus_rows.append(row)
    return pd.DataFrame(assay_rows), locus_rows


def match_locus(row: pd.Series, loci_by_assay: dict[str, list[dict]]) -> tuple[dict | None, str]:
    candidates = [locus for locus in loci_by_assay.get(str(row["assay_target_ID"]), []) if row["SNP"] in locus["rs_aliases"]]
    if not candidates:
        return None, "unresolved_no_exact_or_reported_LD_class_match"
    candidates.sort(key=lambda item: float(item.get("P-value_joint_model") or 1.0))
    if len(candidates) == 1:
        return candidates[0], "matched_source_pQTL_locus"
    return candidates[0], "multiple_source_loci_best_joint_P_selected"


def build_pav_audit(instruments: pd.DataFrame, assay_annotations: pd.DataFrame, locus_rows: list[dict]) -> pd.DataFrame:
    loci_by_assay: dict[str, list[dict]] = defaultdict(list)
    for locus in locus_rows:
        loci_by_assay[str(locus["SeqId"])].append(locus)
    assay_map = assay_annotations.set_index("SeqId").to_dict("index")
    records = []
    for _, instrument in instruments.iterrows():
        locus, status = match_locus(instrument, loci_by_assay)
        assay = assay_map.get(str(instrument["assay_target_ID"]), {})
        same_gene = truthy(locus.get("any_coding_same_gene")) if locus else None
        diff_gene = truthy(locus.get("any_coding_diff_gene")) if locus else None
        assay_pav = truthy(assay.get("PAV_soma"))
        if same_gene is True:
            risk = "target_gene_PAV_in_source_LD_class"
        elif locus is None:
            risk = "unresolved_no_source_locus_match"
        elif assay_pav is True:
            risk = "assay_level_PAV_reported_but_not_linked_to_this_instrument"
        else:
            risk = "no_target_gene_PAV_in_matched_source_record"
        records.append({
            "gene_symbol": instrument["gene_symbol"],
            "assay_target_ID": instrument["assay_target_ID"],
            "UniProt_ID": instrument["UniProt_ID"],
            "SNP": instrument["SNP"],
            "chromosome_hg19": instrument["chromosome"],
            "position_hg19": instrument["position_hg19_from_liftover"],
            "cis_to_encoding_gene": instrument["cis_to_encoding_gene"],
            "F_statistic": instrument["F_statistic"],
            "source_locus_match_status": status,
            "source_pQTL_ID_global": locus.get("pQTL_ID_global") if locus else "NA",
            "source_variant": locus.get("variant") if locus else "NA",
            "source_cis_trans": locus.get("cis_trans") if locus else "NA",
            "source_any_coding_same_gene": same_gene,
            "source_any_coding_diff_gene": diff_gene,
            "source_cis_eqtl_coding_genes": locus.get("cis_eqtl_coding_genes") if locus else "NA",
            "source_joint_P": locus.get("P-value_joint_model") if locus else "NA",
            "assay_PAV_soma": assay_pav,
            "assay_cis_pQTL_soma": truthy(assay.get("cis_pQTL_soma")),
            "assay_eQTL_soma": truthy(assay.get("eQTL_soma")),
            "assay_confidence_tier": assay.get("confidence_tier", "NA"),
            "PAV_epitope_risk_classification": risk,
            "automatic_exclusion": False,
            "audit_rule": "PAV annotation is mandatory but unresolved status is not treated as absence or automatic exclusion",
        })
    return pd.DataFrame(records)


def run_ld_audit(instruments: pd.DataFrame) -> pd.DataFrame:
    WORK_DIR.mkdir(parents=True, exist_ok=True)
    snp_file = WORK_DIR / "decode_smp_instrument_snps.txt"
    snp_file.write_text("\n".join(sorted(instruments["SNP"].unique())) + "\n", encoding="utf-8")
    prefix = WORK_DIR / "decode_smp_instrument_pairs_r2_0_8"
    ld_file = Path(str(prefix) + ".ld.gz")
    command = [
        str(PLINK), "--bfile", str(LD_REFERENCE), "--extract", str(snp_file),
        "--r2", "gz", "--ld-window", "99999", "--ld-window-kb", "1000000",
        "--ld-window-r2", "0.8", "--out", str(prefix),
    ]
    completed = subprocess.run(command, text=True, capture_output=True, check=False)
    (WORK_DIR / "plink_console.log").write_text(completed.stdout + "\n" + completed.stderr, encoding="utf-8")
    if completed.returncode != 0 or not ld_file.exists():
        raise RuntimeError("PLINK cross-assay LD audit failed")

    assay_by_snp = instruments.groupby("SNP")["assay_target_ID"].apply(lambda x: sorted(set(x))).to_dict()
    gene_by_snp = instruments.groupby("SNP")["gene_symbol"].apply(lambda x: sorted(set(x))).to_dict()
    records = []
    for snp, assays in assay_by_snp.items():
        if len(assays) > 1:
            records.append({
                "row_type": "exact_shared_SNP", "SNP_A": snp, "SNP_B": snp, "R2": 1.0,
                "assays_A": ";".join(assays), "assays_B": ";".join(assays),
                "genes_A": ";".join(gene_by_snp[snp]), "genes_B": ";".join(gene_by_snp[snp]),
                "cross_assay_reuse": True,
            })
    if ld_file.stat().st_size > 0:
        ld = pd.read_csv(ld_file, sep=r"\s+", compression="gzip")
        for row in ld.itertuples(index=False):
            assays_a = assay_by_snp.get(row.SNP_A, [])
            assays_b = assay_by_snp.get(row.SNP_B, [])
            if not assays_a or not assays_b or not set(assays_a).isdisjoint(assays_b):
                continue
            records.append({
                "row_type": "distinct_SNP_pair_in_high_LD", "SNP_A": row.SNP_A, "SNP_B": row.SNP_B,
                "R2": row.R2, "assays_A": ";".join(assays_a), "assays_B": ";".join(assays_b),
                "genes_A": ";".join(gene_by_snp[row.SNP_A]), "genes_B": ";".join(gene_by_snp[row.SNP_B]),
                "cross_assay_reuse": True,
            })
    if not records:
        return pd.DataFrame(columns=["row_type", "SNP_A", "SNP_B", "R2", "assays_A", "assays_B", "genes_A", "genes_B", "cross_assay_reuse"])
    return pd.DataFrame(records).drop_duplicates().sort_values(["row_type", "R2", "SNP_A", "SNP_B"], ascending=[True, False, True, True])


def main() -> None:
    for required in (WORKBOOK, INSTRUMENTS, PLINK, Path(str(LD_REFERENCE) + ".bed")):
        if not required.exists():
            raise FileNotFoundError(required)
    instruments = pd.read_csv(INSTRUMENTS, sep="\t")
    if instruments["assay_target_ID"].nunique() != 9:
        raise ValueError("Expected 9 SMP assays in the genome-wide instrument table")
    assay_annotations, locus_rows = read_target_supplement(set(instruments["assay_target_ID"].astype(str)))
    if len(assay_annotations) != 9:
        raise ValueError(f"Expected 9 exact assay annotations, found {len(assay_annotations)}")

    pav = build_pav_audit(instruments, assay_annotations, locus_rows)
    shared = run_ld_audit(instruments)
    assay_columns = [
        "UniProt", "OlinkID", "SeqId", "Gene", "Targeted_Olink", "Targeted_Soma",
        "cis_pQTL_soma", "PAV_soma", "eQTL_soma", "location", "N_platforms_tested",
        "N_platforms_with_cis_pQTL", "confidence_tier",
    ]
    assay_annotations[[column for column in assay_columns if column in assay_annotations.columns]].to_csv(
        ASSAY_OUTPUT, sep="\t", index=False, na_rep="NA"
    )
    pav.to_csv(PAV_OUTPUT, sep="\t", index=False, na_rep="NA")
    shared.to_csv(SHARED_OUTPUT, sep="\t", index=False, na_rep="NA")
    LOG_OUTPUT.write_text("\n".join([
        f"instrument_rows={len(instruments)}",
        f"assay_annotation_rows={len(assay_annotations)}",
        f"source_pqtl_locus_rows_for_targets={len(locus_rows)}",
        f"instrument_source_locus_matches={(pav['source_locus_match_status'] != 'unresolved_no_exact_or_reported_LD_class_match').sum()}",
        f"target_gene_PAV_linked={(pav['PAV_epitope_risk_classification'] == 'target_gene_PAV_in_source_LD_class').sum()}",
        f"cross_assay_shared_or_high_LD_rows={len(shared)}",
        f"workbook_sha256={sha256(WORKBOOK)}",
        "PAV_unresolved_not_treated_as_no_risk=true",
        "PAV_not_used_as_automatic_exclusion=true",
    ]) + "\n", encoding="utf-8")
    print(LOG_OUTPUT.read_text(encoding="utf-8"))


if __name__ == "__main__":
    main()
