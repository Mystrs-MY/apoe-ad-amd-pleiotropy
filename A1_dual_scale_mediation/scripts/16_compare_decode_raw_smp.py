#!/usr/bin/env python3
"""Compute gated raw mediation and compare it with the primary SMP extension."""

from __future__ import annotations

import math
from pathlib import Path

import numpy as np
import pandas as pd
import yaml
from openpyxl import load_workbook
from scipy.stats import norm


ROOT = Path(__file__).resolve().parents[1]
PREFIX = "decode_raw_gated_10074_128_8687_26"
ALPHA_SOURCE = ROOT / "data_processed" / f"{PREFIX}_direct_APOE_alpha.tsv"
TOTAL_CONFIG = ROOT.parent / "A1_protein_upgrade" / "config" / "outcomes.yml"
WORKBOOK = ROOT / "data_raw" / "decode_public" / "41586_2023_6563_MOESM3_ESM.xlsx"
RAW_CIS_IV = ROOT / "data_processed" / f"{PREFIX}_clumped_instruments_cis_only.tsv"

ALPHA_OUTPUT = ROOT / "tables" / f"APOE_variant_to_{PREFIX}_alpha.tsv"
MEDIATION_OUTPUT = ROOT / "tables" / f"{PREFIX}_two_step_mediation.tsv"
MEDIATION_CIS_OUTPUT = ROOT / "tables" / f"{PREFIX}_two_step_mediation_cis_only.tsv"
MEDIATION_CIS_PAV_OUTPUT = ROOT / "tables" / f"{PREFIX}_two_step_mediation_cis_only_PAV_filtered.tsv"
PAV_OUTPUT = ROOT / "tables" / f"{PREFIX}_PAV_audit.tsv"
COMPARISON_OUTPUT = ROOT / "tables" / "decode_raw_vs_SMP_normalization_comparison.tsv"
SUMMARY_OUTPUT = ROOT / "tables" / "decode_raw_gated_sensitivity_summary.tsv"

VARIANTS = {"rs429358": ("C", "T", "rs429358_C"), "rs7412": ("T", "C", "rs7412_T")}
OUTCOMES = ["AD", "any_AMD", "dry_AMD", "wet_AMD"]
PLANNED_PATHS = 2 * 2 * 4
BOOTSTRAP_DRAWS = 100_000
SEED = 20260719 + 10


def bh(values: pd.Series, n: int) -> pd.Series:
    out = pd.Series(np.nan, index=values.index)
    valid = values.dropna().sort_values()
    if valid.empty:
        return out
    adjusted = valid.to_numpy() * n / np.arange(1, len(valid) + 1)
    adjusted = np.minimum.accumulate(adjusted[::-1])[::-1]
    out.loc[valid.index] = np.minimum(1, adjusted)
    return out


def build_alpha() -> pd.DataFrame:
    alpha = pd.read_csv(ALPHA_SOURCE, sep="\t").rename(columns={
        "SomaScan_SeqId": "assay_target_ID", "requested_variant": "variant",
        "Beta": "alpha", "SE": "alpha_SE", "Pval": "alpha_P", "N": "alpha_N",
        "effectAllele": "effect_allele", "otherAllele": "other_allele",
        "ImpMAF": "imputation_MAF",
    })
    if len(alpha) != 4 or alpha.duplicated(["assay_target_ID", "variant"]).any():
        raise ValueError("Expected four unique raw direct-alpha rows")
    for variant, (effect, other, _) in VARIANTS.items():
        rows = alpha[alpha["variant"].eq(variant)]
        if not (rows["effect_allele"].eq(effect) & rows["other_allele"].eq(other)).all():
            raise ValueError(f"Raw alpha allele mismatch for {variant}")
    alpha["normalization"] = "raw_non_normalized"
    alpha["direct_variant_not_proxy"] = True
    alpha["imputation_MAF_not_EAF"] = True
    alpha["alpha_selection_role"] = "not_used_for_gate_or_eligibility"
    alpha["inference_role"] = "outcome_triggered_normalization_robustness_not_independent_confirmation"
    return alpha


def annotate_raw_cis_pav() -> pd.DataFrame:
    instruments = pd.read_csv(RAW_CIS_IV, sep="\t")
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    sheet = workbook["ST20_somascan_pQTLs_nonnorm"]
    headers = [str(v).replace("\n", "_").replace(" ", "_") if v is not None else ""
               for v in next(sheet.iter_rows(min_row=4, max_row=4, values_only=True))]
    loci = []
    target_ids = set(instruments["assay_target_ID"].astype(str))
    for values in sheet.iter_rows(min_row=5, values_only=True):
        row = dict(zip(headers, values))
        if str(row.get("SeqId")) in target_ids:
            aliases = {item.strip() for value in (row.get("variant"), row.get("LD_class"))
                       if value is not None for item in str(value).split(",") if item.strip().startswith("rs")}
            row["aliases"] = aliases
            loci.append(row)
    records = []
    for row in instruments.itertuples(index=False):
        matches = [locus for locus in loci if str(locus.get("SeqId")) == str(row.assay_target_ID) and row.SNP in locus["aliases"]]
        match = matches[0] if len(matches) == 1 else None
        same_gene_pav = str(match.get("any_coding_same_gene")).upper() == "Y" if match else None
        records.append({
            "gene_symbol": row.gene_symbol, "assay_target_ID": row.assay_target_ID, "SNP": row.SNP,
            "source_ST20_match_status": "matched_unique_LD_class" if match else "unresolved_or_multiple",
            "source_variant": match.get("variant") if match else "NA",
            "source_cis_trans": match.get("cis_trans") if match else "NA",
            "source_any_coding_same_gene": same_gene_pav,
            "source_cis_eqtl_coding_genes": match.get("cis_eqtl_coding_genes") if match else "NA",
            "target_gene_PAV_linked": same_gene_pav is True,
            "automatic_exclusion": False,
            "audit_boundary": "PAV unresolved is not treated as no risk; filtering is a separate sensitivity",
        })
    return pd.DataFrame(records)


def mediation(alpha: pd.DataFrame, beta_file: Path, scope: str, total: dict,
              output: Path, pav_remove: set[tuple[str, str]] | None = None) -> pd.DataFrame:
    beta = pd.read_csv(beta_file, sep="\t")
    beta = beta[(beta["method_role"].eq("primary"))].copy()
    beta["exclusion_reason"] = beta["exclusion_reason"].astype("object")
    if pav_remove:
        harmonized_path = ROOT / "data_processed" / f"{PREFIX}_harmonized_data_cis_only.tsv.gz"
        harmonized = pd.read_csv(harmonized_path, sep="\t", compression="gzip")
        affected = {assay for assay, _ in pav_remove}
        for assay in affected:
            remaining = harmonized[(harmonized["assay_target_ID"].eq(assay)) & harmonized["mr_keep"].eq(True)
                                   & ~harmonized["SNP"].isin([snp for a, snp in pav_remove if a == assay])]
            if remaining.empty:
                beta.loc[beta["assay_target_ID"].eq(assay), ["beta", "SE", "P_value"]] = np.nan
                beta.loc[beta["assay_target_ID"].eq(assay), "beta_status"] = "not_estimable"
                beta.loc[beta["assay_target_ID"].eq(assay), "exclusion_reason"] = "All raw cis instruments removed by target-gene PAV filter"

    grid = pd.MultiIndex.from_product([
        sorted(alpha["assay_target_ID"].unique()), VARIANTS, OUTCOMES
    ], names=["assay_target_ID", "variant", "outcome"]).to_frame(index=False)
    joined = grid.merge(alpha[["gene_symbol", "assay_target_ID", "variant", "alpha", "alpha_SE", "alpha_P"]],
                        on=["assay_target_ID", "variant"], how="left")
    joined = joined.merge(beta[["assay_target_ID", "outcome", "method", "nsnp", "beta", "SE", "P_value",
                               "beta_status", "Q_P", "min_F"]].rename(columns={
                                   "beta": "protein_outcome_beta", "SE": "protein_outcome_SE",
                                   "P_value": "protein_outcome_P", "beta_status": "protein_outcome_beta_status"
                               }), on=["assay_target_ID", "outcome"], how="left")
    joined["analysis_scope"] = scope
    joined["planned_paths"] = PLANNED_PATHS
    joined["inference_role"] = "outcome_triggered_normalization_robustness_not_independent_confirmation"
    joined["total_effect"] = [total[VARIANTS[v][2]][o]["beta"] for v, o in zip(joined.variant, joined.outcome)]
    rng = np.random.default_rng(SEED + {"genome_wide": 0, "cis_only": 1, "cis_only_PAV_filtered": 2}[scope])
    rows = []
    for _, row in joined.iterrows():
        rec = row.to_dict()
        if row.get("protein_outcome_beta_status") != "reestimated" or pd.isna(row.get("protein_outcome_beta")):
            rec.update({"mediation_status": "not_estimable", "indirect_effect": np.nan,
                        "delta_SE": np.nan, "delta_P": np.nan, "delta_P_display": "NA",
                        "mediated_proportion": np.nan, "bootstrap_CI_lower": np.nan,
                        "bootstrap_CI_upper": np.nan, "direction_classification": "not_estimable"})
        else:
            indirect = row.alpha * row.protein_outcome_beta
            se = math.sqrt((row.protein_outcome_beta * row.alpha_SE) ** 2 +
                           (row.alpha * row.protein_outcome_SE) ** 2)
            logp = math.log(2) + norm.logsf(abs(indirect / se))
            p = math.exp(logp) if logp > math.log(np.finfo(float).tiny) else 0.0
            draws = rng.normal(row.alpha, row.alpha_SE, BOOTSTRAP_DRAWS) * rng.normal(
                row.protein_outcome_beta, row.protein_outcome_SE, BOOTSTRAP_DRAWS)
            rec.update({
                "mediation_status": "estimable", "indirect_effect": indirect, "delta_SE": se,
                "delta_P": p, "delta_P_display": "<1e-300" if -logp / math.log(10) > 300 else f"{p:.3e}",
                "mediated_proportion": indirect / row.total_effect,
                "bootstrap_CI_lower": np.quantile(draws, .025), "bootstrap_CI_upper": np.quantile(draws, .975),
                "direction_classification": "concordant_mediation" if np.sign(indirect) == np.sign(row.total_effect)
                else "opposing_or_suppressing",
            })
        rows.append(rec)
    result = pd.DataFrame(rows)
    estimable = result["mediation_status"].eq("estimable")
    result["P_FDR_16"] = np.nan
    result.loc[estimable, "P_FDR_16"] = bh(result.loc[estimable, "delta_P"], PLANNED_PATHS)
    result["P_Bonferroni_16"] = np.where(estimable, np.minimum(1, result["delta_P"] * PLANNED_PATHS), np.nan)
    result["passes_Bonferroni_16"] = result["P_Bonferroni_16"].lt(.05)
    result["bootstrap_draws"] = BOOTSTRAP_DRAWS
    result["selection_warning"] = "raw assays selected after SMP outcome results; corrected P values are descriptive robustness metrics"
    result.to_csv(output, sep="\t", index=False, na_rep="NA")
    return result


def main() -> None:
    with TOTAL_CONFIG.open("rt", encoding="utf-8") as handle:
        total = yaml.safe_load(handle)["apoe_total_effects"]
    alpha = build_alpha()
    alpha.to_csv(ALPHA_OUTPUT, sep="\t", index=False, na_rep="NA")
    pav = annotate_raw_cis_pav()
    pav.to_csv(PAV_OUTPUT, sep="\t", index=False, na_rep="NA")
    remove = {(r.assay_target_ID, r.SNP) for r in pav[pav["target_gene_PAV_linked"]].itertuples()}
    raw_gw = mediation(alpha, ROOT / "tables" / f"{PREFIX}_beta_results.tsv", "genome_wide", total, MEDIATION_OUTPUT)
    raw_cis = mediation(alpha, ROOT / "tables" / f"{PREFIX}_beta_results_cis_only.tsv", "cis_only", total, MEDIATION_CIS_OUTPUT)
    raw_pav = mediation(alpha, ROOT / "tables" / f"{PREFIX}_beta_results_cis_only.tsv", "cis_only_PAV_filtered",
                        total, MEDIATION_CIS_PAV_OUTPUT, remove)

    smp = pd.read_csv(ROOT / "tables" / "decode_smp_two_step_mediation_cis_only.tsv", sep="\t")
    comparison = smp[smp["assay_target_ID"].isin(alpha["assay_target_ID"])][[
        "gene_symbol", "assay_target_ID", "variant", "outcome", "alpha", "protein_outcome_beta",
        "indirect_effect", "mediation_P_Bonferroni_72", "direction_classification"
    ]].rename(columns={column: f"SMP_{column}" for column in ["alpha", "protein_outcome_beta", "indirect_effect",
                                                                "mediation_P_Bonferroni_72", "direction_classification"]})
    comparison = comparison.merge(raw_cis[[
        "assay_target_ID", "variant", "outcome", "alpha", "protein_outcome_beta", "indirect_effect",
        "P_Bonferroni_16", "direction_classification"
    ]].rename(columns={column: f"raw_{column}" for column in ["alpha", "protein_outcome_beta", "indirect_effect",
                                                                "P_Bonferroni_16", "direction_classification"]}),
        on=["assay_target_ID", "variant", "outcome"], how="left")
    comparison["alpha_direction_consistent"] = np.sign(comparison["SMP_alpha"]) == np.sign(comparison["raw_alpha"])
    comparison["beta_direction_consistent"] = np.sign(comparison["SMP_protein_outcome_beta"]) == np.sign(comparison["raw_protein_outcome_beta"])
    comparison["indirect_direction_consistent"] = np.sign(comparison["SMP_indirect_effect"]) == np.sign(comparison["raw_indirect_effect"])
    comparison["comparison_role"] = "normalization_robustness_only"
    comparison.to_csv(COMPARISON_OUTPUT, sep="\t", index=False, na_rep="NA")

    summary = pd.DataFrame([
        {"scope": "raw_genome_wide", "estimable_paths": (raw_gw.mediation_status == "estimable").sum(),
         "Bonferroni_16_survivors": raw_gw.passes_Bonferroni_16.sum()},
        {"scope": "raw_cis_only", "estimable_paths": (raw_cis.mediation_status == "estimable").sum(),
         "Bonferroni_16_survivors": raw_cis.passes_Bonferroni_16.sum()},
        {"scope": "raw_cis_only_PAV_filtered", "estimable_paths": (raw_pav.mediation_status == "estimable").sum(),
         "Bonferroni_16_survivors": raw_pav.passes_Bonferroni_16.sum()},
    ])
    summary["independent_confirmation"] = False
    summary["reason"] = "two assays were selected after SMP outcome results"
    summary.to_csv(SUMMARY_OUTPUT, sep="\t", index=False)
    print(summary.to_string(index=False))
    print(pav.to_string(index=False))


if __name__ == "__main__":
    main()
